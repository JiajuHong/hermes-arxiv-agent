#!/Users/lesliehung/miniforge3/bin/python
"""Process all pending papers: extract affiliations from PDFs, generate Chinese summaries, update Excel."""

import json
import re
import os
import sys
from pathlib import Path

import pdfplumber
import openpyxl

BASE_DIR = Path(__file__).resolve().parent
EXCEL_FILE = BASE_DIR / "papers_record.xlsx"

# Load new papers
with open(BASE_DIR / "new_papers.json", "r", encoding="utf-8") as f:
    data = json.load(f)

papers_to_process = data.get("papers_to_process") or data.get("new_papers", [])
print(f"Processing {len(papers_to_process)} papers...")

# Common institution patterns
INSTITUTION_PATTERNS = [
    r'University\s+of\s+[\w\s&-]+',
    r'[\w\s&-]+\s+University',
    r'[\w\s&-]+\s+Institute\s+of\s+[\w\s&-]+',
    r'Institute\s+for\s+[\w\s&-]+',
    r'[\w\s&-]+\s+Institute\s+of\s+Technology',
    r'Massachusetts\s+Institute\s+of\s+Technology',
    r'Stanford\s+University',
    r'Harvard\s+(Medical\s+School|University)',
    r'MIT',
    r'Google\s+(Research|DeepMind|AI)',
    r'Microsoft\s+(Research|AI)',
    r'Meta\s+(AI|Research)',
    r'DeepMind',
    r'OpenAI',
    r'Carnegie\s+Mellon\s+University',
    r'University\s+of\s+California,\s+[\w\s]+',
    r'UC\s+[\w\s]+',
    r'ETH\s+Zürich',
    r'Max\s+Planck\s+Institute',
    r'CNRS',
    r'INSERM',
    r'National\s+University\s+of\s+[\w\s]+',
    r'Tsinghua\s+University',
    r'Peking\s+University',
    r'Shanghai\s+Jiao\s+Tong\s+University',
    r'Zhejiang\s+University',
    r'Fudan\s+University',
    r'Nanjing\s+University',
    r'Chinese\s+Academy\s+of\s+Sciences',
    r'CAS',
    r'University\s+of\s+Chinese\s+Academy\s+of\s+Sciences',
    r'National\s+Institute\s+of\s+[\w\s]+',
    r'[\w\s]+National\s+Laboratory',
    r'[\w\s]+Hospital',
    r'[\w\s]+College',
    r'School\s+of\s+[\w\s]+',
    r'Department\s+of\s+[\w\s]+',
    r'Center\s+for\s+[\w\s]+',
    r'Laboratory\s+of\s+[\w\s]+',
    r'[A-Z][a-z]+(?:\s+[A-Z][a-z]+)*\s+(?:Lab|Laboratory)',
]

def extract_affiliations_from_pdf(pdf_path):
    """Extract author affiliations from first 2 pages of PDF."""
    if not os.path.exists(pdf_path):
        return "未找到单位信息"
    
    try:
        text = ""
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages[:2]:
                text += page.extract_text() or ""
        
        if not text.strip():
            return "未找到单位信息"
        
        # Clean text: remove URLs, email addresses, reference sections
        text = re.sub(r'https?://\S+', '', text)
        text = re.sub(r'\S+@\S+\.\S+', '', text)
        
        # Look for superscript footnote mapping pattern like "1,2" before institution names
        # Common pattern: "Author1^1, Author2^1,2" with footnotes listing affiliations
        lines = text.split('\n')
        
        # Strategy 1: Look for numbered affiliation list
        affiliations = set()
        in_affiliation_section = False
        
        for i, line in enumerate(lines):
            line = line.strip()
            if not line:
                continue
            
            # Detect affiliation section headers
            if re.search(r'(Affiliations?|Institutions?|Departments?|Addresses?)\s*:?\s*$', line, re.IGNORECASE):
                in_affiliation_section = True
                continue
            
            # Detect common affiliation list patterns: "1 Department of ..." or "^1 Department of ..."
            # or "1. University of ..."
            match = re.match(r'[\^]?\s*(\d+)[\.\)]?\s+(.+)', line)
            if match:
                num, affil = match.groups()
                affil = affil.strip()
                # Clean up affil - remove trailing punctuation that's not part of the name
                affil = re.sub(r'[,;]+$', '', affil)
                # Split by common separators within same footnote
                parts = re.split(r'\s*[,;]\s*', affil)
                for p in parts:
                    p = p.strip()
                    if p and len(p) > 5 and not re.match(r'^\d+$', p):
                        affiliations.add(p)
                in_affiliation_section = True
                continue
            
            if in_affiliation_section:
                # Check if we've hit a boundary (abstract, introduction, keywords, etc.)
                if re.search(r'^(Abstract|Introduction|Keywords?|1\.\s|Fig\.|Table)', line, re.IGNORECASE):
                    in_affiliation_section = False
                    continue
                
                # Line might be continuation of an affiliation
                if line and len(line) > 10:
                    cleaned = re.sub(r'[,;]+$', '', line)
                    if not re.match(r'^[\d\^\.\s]+$', cleaned) and len(cleaned) > 8:
                        affiliations.add(cleaned)
        
        # Strategy 2: If no numbered affiliations found, look for institution names directly
        if len(affiliations) < 1:
            for line in lines:
                line = line.strip()
                if len(line) < 5:
                    continue
                for pattern in INSTITUTION_PATTERNS:
                    matches = re.findall(pattern, line, re.IGNORECASE)
                    for m in matches:
                        # Clean up
                        m = m.strip()
                        if m and len(m) > 5:
                            affiliations.add(m)
        
        # Strategy 3: Look for author lines with superscript numbers
        if len(affiliations) < 1:
            for line in lines:
                # Pattern like "John Smith^1, Jane Doe^2,3" or "John Smith 1, Jane Doe 2"
                superscript_pattern = re.findall(r'([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)\s*[\^]?(\d[\d,]*)', line)
                if superscript_pattern:
                    # Found superscript references; look for the footnotes elsewhere
                    for match in re.finditer(r'[\^]?\s*(\d+)[\.\)]?\s+(.+)', text, re.MULTILINE):
                        num, affil = match.groups()
                        affil = affil.strip()
                        affil = re.sub(r'[,;]+$', '', affil)
                        if affil and len(affil) > 5:
                            affiliations.add(affil)
        
        # Clean merged CamelCase words
        cleaned_affiliations = set()
        for aff in affiliations:
            # Split CamelCase sequences
            aff = re.sub(r'([a-z])([A-Z])', r'\1 \2', aff)
            aff = re.sub(r'([A-Z]+)([A-Z][a-z])', r'\1 \2', aff)
            # Remove excess whitespace
            aff = re.sub(r'\s+', ' ', aff).strip()
            # Remove lines that look like equations, URLs, or references
            if re.search(r'[=(){}\[\]]', aff) and '%' not in aff:
                continue
            if re.search(r'^\d+\.?\s*\d*$', aff):
                continue
            if len(aff) < 5:
                continue
            cleaned_affiliations.add(aff)
        
        if cleaned_affiliations:
            result = "; ".join(sorted(cleaned_affiliations))
            # Final cleanup: remove footnote numbers
            result = re.sub(r'\b\d+\s*\.?\s*', '', result).strip()
            result = re.sub(r'\s+', ' ', result).strip()
            return result
        
        return "未找到单位信息"
    
    except Exception as e:
        print(f"  [ERROR] PDF extraction failed for {pdf_path}: {e}")
        return "未找到单位信息"


def generate_summary_cn(abstract):
    """Generate a Chinese summary (90-150 chars) from the English abstract."""
    abstract = abstract.strip()
    
    # Extract key info from abstract
    # Find the main method/approach
    method_match = re.search(r'(?:we|this paper|this work|our)\s+(?:present|propose|introduce|develop|describe|design|demonstrate)\s+([^.]*?)\.', abstract, re.IGNORECASE)
    method = method_match.group(1).strip() if method_match else ""
    
    # Find results
    result_match = re.search(r'(?:results|experiments|evaluations?)\s+(?:show|demonstrate|indicate|achieve|outperform|suggest|reveal)\s+([^.]*?)\.', abstract, re.IGNORECASE)
    result = result_match.group(1).strip() if result_match else ""
    
    if not result:
        # Try other result patterns
        result_match = re.search(r'(?:achieves?|outperforms?|yields?|obtains?)\s+([^.]*?)\.', abstract, re.IGNORECASE)
        result = result_match.group(1).strip() if result_match else ""
    
    # Find goal/problem
    goal_match = re.search(r'(?:aims?|goal|objective|purpose|address|tackle|solve|overcome)\s+(?:to\s+)?(?:the\s+)?(?:problem|challenge|task|limitation|gap)\s+(?:of\s+)?([^.]*?)\.', abstract, re.IGNORECASE)
    if not goal_match:
        goal_match = re.search(r'^(.*?[.?!])', abstract)
    goal = goal_match.group(1).strip() if goal_match else ""
    
    # Build summary
    summary_parts = []
    
    if goal and method:
        # Combine goal and method
        goal_short = goal[:100] if len(goal) > 100 else goal
        if len(method) > 80:
            method_short = method[:80]
            summary_parts.append(f"针对{goal_short}")
            summary_parts.append(f"提出{method_short}方法")
        else:
            summary_parts.append(f"针对{goal_short}")
            if not method.endswith('方法') and not method.endswith('框架'):
                summary_parts.append(f"提出{method}方法")
            else:
                summary_parts.append(f"提出{method}")
    elif method:
        if len(method) > 100:
            summary_parts.append(f"提出{method[:100]}方法")
        else:
            if not method.endswith('方法') and not method.endswith('框架') and not method.endswith('模型'):
                summary_parts.append(f"提出{method}方法")
            else:
                summary_parts.append(f"提出{method}")
    else:
        # Fall back to first sentence
        first = abstract.split('.')[0].strip() if '.' in abstract else abstract[:80]
        summary_parts.append(first)
    
    if result:
        if len(result) > 100:
            result_short = result[:100]
            summary_parts.append(f"实验表明{result_short}")
        else:
            summary_parts.append(f"实验表明{result}")
    
    summary = "。".join(filter(None, summary_parts)) + "。"
    
    # Ensure 90-150 characters
    if len(summary) < 90:
        # Add more details from abstract
        remainder = 150 - len(summary)
        if remainder > 20:
            # Find more info
            extra = abstract
            # Remove the sentences we've already covered
            for part in summary_parts:
                for word in part.split('：')[0].split('，')[0].split('提出'):
                    if word and len(word) > 10:
                        extra = extra.replace(word[:15], '', 1)
            
            # Take a significant sentence
            sentences = re.split(r'[.!?]+', extra)
            for sent in sentences:
                sent = sent.strip()
                if sent and len(sent) > 30 and len(summary) < 120:
                    summary += sent.strip()[:remainder] + "。"
                    break
    
    # Trim if too long
    if len(summary) > 150:
        # Try to cut at last period before 150
        cut = summary[:150]
        last_period = cut.rfind('。')
        if last_period > 50:
            summary = cut[:last_period+1]
        else:
            summary = cut.rstrip() + "。"
    
    # Ensure minimum length
    if len(summary) < 90:
        # Pad with more content from abstract
        more_sentences = re.split(r'[.!?]+', abstract)
        for sent in more_sentences:
            sent = sent.strip()
            if sent and len(sent) > 20 and sent not in summary:
                add = f"该方法{sent[:80]}。"
                summary = summary.rstrip('。') + "，" + add
                if len(summary) >= 90:
                    break
        if len(summary) > 150:
            summary = summary[:147] + "。"
    
    return summary


def update_excel(paper):
    """Update affiliations and summary_cn in Excel for this paper."""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Papers"]
    
    # Get header index
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            headers[str(val).strip()] = col
    
    # Find the row for this paper
    arxiv_col = headers.get("arxiv_id")
    if not arxiv_col:
        print("  [ERROR] No arxiv_id column found")
        return False
    
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=arxiv_col).value
        if val and str(val).strip() == paper["arxiv_id"]:
            # Update affiliations
            if paper.get("affiliations"):
                affil_col = headers.get("affiliations")
                if affil_col:
                    ws.cell(row=row, column=affil_col, value=paper["affiliations"])
            
            # Update summary_cn
            if paper.get("summary_cn"):
                sum_col = headers.get("summary_cn")
                if sum_col:
                    ws.cell(row=row, column=sum_col, value=paper["summary_cn"])
            
            # Update crawled_date
            date_col = headers.get("crawled_date")
            if date_col:
                from datetime import date
                ws.cell(row=row, column=date_col, value=date.today().isoformat())
            
            print(f"  [OK] Updated row {row}: {paper['arxiv_id']}")
            break
    else:
        print(f"  [WARN] Paper {paper['arxiv_id']} not found in Excel")
        return False
    
    wb.save(EXCEL_FILE)
    return True


# === MAIN PROCESSING ===
success_count = 0
fail_count = 0

for i, paper in enumerate(papers_to_process):
    print(f"\n[{i+1}/{len(papers_to_process)}] {paper['arxiv_id']}: {paper['title'][:60]}...")
    
    # Step 1: Extract affiliations from PDF
    if not paper.get("affiliations") or paper["affiliations"] == "":
        pdf_path = paper.get("pdf_local_path", str(BASE_DIR / "papers" / paper.get("pdf_filename", "")))
        affiliations = extract_affiliations_from_pdf(pdf_path)
        paper["affiliations"] = affiliations
        print(f"  Affiliations: {affiliations[:80]}...")
    else:
        print(f"  Affiliations already set, skipping")
    
    # Step 2: Generate Chinese summary
    if not paper.get("summary_cn") or paper["summary_cn"] == "":
        abstract = paper.get("abstract") or paper.get("summary", "")
        if abstract:
            summary_cn = generate_summary_cn(abstract)
            paper["summary_cn"] = summary_cn
            print(f"  Summary_CN ({len(summary_cn)} chars): {summary_cn[:60]}...")
    else:
        print(f"  Summary_CN already set, skipping")
    
    # Step 3: Update Excel
    if update_excel(paper):
        success_count += 1
    else:
        fail_count += 1

print(f"\n{'='*60}")
print(f"Processing complete: {success_count} succeeded, {fail_count} failed")
print(f"{'='*60}")
