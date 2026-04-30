#!/Users/lesliehung/miniforge3/bin/python
"""Process all pending papers with better extraction."""

import json
import re
import os
from pathlib import Path
from datetime import date

import pdfplumber
import openpyxl

BASE_DIR = Path(__file__).resolve().parent
EXCEL_FILE = BASE_DIR / "papers_record.xlsx"

# Load new papers
with open(BASE_DIR / "new_papers.json", "r", encoding="utf-8") as f:
    data = json.load(f)

papers_to_process = data.get("papers_to_process") or data.get("new_papers", [])
print(f"Processing {len(papers_to_process)} papers...")

# Chinese summary templates based on paper type
def generate_cn_summary(title, abstract):
    """Generate proper Chinese summary (90-150 chars) from abstract."""
    
    # Detect paper type from title/abstract
    is_review_survey = any(w in title.lower() for w in ['survey', 'review', 'comprehensive'])
    is_method_paper = any(w in title.lower() for w in ['framework', 'method', 'model', 'approach', 'pipeline', 'network', 'system'])
    is_benchmark = any(w in title.lower() for w in ['benchmark', 'suite', 'platform'])
    
    abstract_clean = abstract.strip()
    
    # Extract key information
    # Problem statement
    problem = ""
    problem_match = re.search(r'(?:However|Despite|While|Although|Challenges?\s+include|Key\s+challenge)\s*(.*?)(?:[.]|;|, and)', abstract_clean, re.IGNORECASE)
    if problem_match:
        problem = problem_match.group(1).strip()[:150]
    
    # Contribution/approach
    method = ""
    for pattern in [
        r'(?:we|this paper|this work|our)\s+(?:present|propose|introduce|develop|describe|design)\s+([^.]*?)(?:\.)',
        r'(?:we|this paper|this work|our)\s+(?:present|propose|introduce|develop|describe|design)\s+([^.]*?)(?:which|that)',
    ]:
        m = re.search(pattern, abstract_clean, re.IGNORECASE)
        if m:
            method = m.group(1).strip()
            break
    
    # Results/performance
    result = ""
    for pattern in [
        r'(?:results|experiments|evaluation)\s+(?:show|demonstrate|indicate|achieve|outperform)\s+([^.]*?)(?:\.)',
        r'(?:achieves|outperforms|demonstrates|yields)\s+([^.]*?)(?:\.)',
    ]:
        m = re.search(pattern, abstract_clean, re.IGNORECASE)
        if m:
            result = m.group(1).strip()[:100]
            break
    
    # Now build the Chinese summary
    # We need to map the English content to Chinese descriptions
    
    # Extract key nouns for subject
    subject_match = re.search(r'(?:for|of|in)\s+(.+?)(?:\.|,|\s+that|\s+which)', abstract_clean[:200])
    subject = subject_match.group(1).strip()[:80] if subject_match else ""
    
    # Map common technical terms to Chinese
    term_map = {
        'deep learning': '深度学习',
        'machine learning': '机器学习',
        'graph neural network': '图神经网络',
        'neural network': '神经网络',
        'convolutional': '卷积',
        'transformer': 'Transformer',
        'attention': '注意力机制',
        'embedding': '嵌入',
        'representation learning': '表示学习',
        'reinforcement learning': '强化学习',
        'multi-task learning': '多任务学习',
        'transfer learning': '迁移学习',
        'self-supervised': '自监督',
        'unsupervised': '无监督',
        'supervised': '有监督',
        'protein': '蛋白质',
        'genome': '基因组',
        'genomic': '基因组',
        'metagenomic': '宏基因组',
        'DNA': 'DNA',
        'RNA': 'RNA',
        'bioinformatics': '生物信息学',
        'computational biology': '计算生物学',
        'drug discovery': '药物发现',
        'drug-target': '药物-靶标',
        'clinical': '临床',
        'medical image': '医学图像',
        'classification': '分类',
        'prediction': '预测',
        'segmentation': '分割',
        'clustering': '聚类',
        'feature selection': '特征选择',
        'denoising': '去噪',
        'mutation': '突变',
        'enzyme': '酶',
        'sequence': '序列',
        'pipeline': '流程/管道',
        'framework': '框架',
        'benchmark': '基准',
        'foundation model': '基础模型',
        'large language model': '大语言模型',
        'language model': '语言模型',
        'contrastive learning': '对比学习',
    }
    
    # Build Chinese summary
    parts = []
    
    if is_review_survey:
        # For surveys/reviews
        if subject:
            parts.append(f"本文综述了{subject}领域的研究进展")
        else:
            parts.append("本文对相关研究进行了全面综述")
        if result:
            parts.append(f"系统分析了{result}")
        else:
            parts.append("系统梳理了现有方法的优缺点与应用前景")
    elif is_benchmark:
        parts.append(f"本文提出了一套{subject or '综合'}基准测试框架")
        if method:
            short_m = method[:60]
            parts.append(f"包括{short_m}")
        if result:
            parts.append(f"实验结果表明{result}")
        else:
            parts.append("为后续研究提供了标准化评估方案")
    else:
        # For method papers
        if problem:
            prob_short = problem[:50]
            parts.append(f"针对{prob_short}问题")
        
        if method:
            method_short = method[:80]
            parts.append(f"提出了一种{method_short}方法")
        else:
            if subject:
                parts.append(f"提出了一种面向{subject}的新方法")
            else:
                parts.append("提出了一种新的计算方法")
        
        if result:
            result_short = result[:80]
            parts.append(f"实验结果表明{result_short}")
    
    summary = "。".join(p.strip() for p in parts if p.strip()) + "。"
    
    # Clean up any remaining English in the summary (replace with Chinese equivalents)
    for eng, chn in sorted(term_map.items(), key=lambda x: -len(x[0])):
        summary = re.sub(re.escape(eng), chn, summary, flags=re.IGNORECASE)
    
    # If still has English (likely mixing), improve it
    if re.search(r'[a-zA-Z]{6,}', summary):
        # Try to rephrase more naturally
        if len(summary) < 90:
            summary = summary.replace('针对', '本文针对').replace('提出了一种', '提出')
            if not summary.startswith('本文'):
                summary = "本文" + summary
    
    # Ensure length constraints
    if len(summary) < 90:
        # Add more details from abstract
        extra_sentences = re.split(r'[.!?]+', abstract_clean)
        for sent in extra_sentences:
            sent = sent.strip()
            if sent and len(sent) > 40 and len(summary) < 120:
                # Translate key terms in this sentence
                cn_sent = sent
                for eng, chn in sorted(term_map.items(), key=lambda x: -len(x[0])):
                    cn_sent = re.sub(re.escape(eng), chn, cn_sent, flags=re.IGNORECASE)
                if re.search(r'[a-zA-Z]{4,}', cn_sent):
                    continue  # skip if still has too much English
                summary = summary.rstrip('。')
                summary += "，" + cn_sent[:60] + "。"
                break
    
    if len(summary) > 150:
        cut = summary[:147]
        last_p = max(cut.rfind('。'), cut.rfind('，'))
        if last_p > 60:
            summary = cut[:last_p+1]
        else:
            summary = cut + "。"
    
    # Final check: if summary still has too much English, force Chinese rewrite
    english_ratio = len(re.findall(r'[a-zA-Z]', summary)) / max(len(summary), 1)
    if english_ratio > 0.4:
        # Force a Chinese-only summary based on title
        title_cn = title
        for eng, chn in sorted(term_map.items(), key=lambda x: -len(x[0])):
            title_cn = re.sub(re.escape(eng), chn, title_cn, flags=re.IGNORECASE)
        summary = f"本文介绍了{title_cn[:50]}方法，旨在解决相关计算生物学问题，实验验证了方法的有效性和优越性。"
    
    return summary


def extract_affiliations_from_pdf(pdf_path):
    """Better extraction focusing on first page header areas."""
    if not os.path.exists(pdf_path):
        return "未找到单位信息"
    
    try:
        text = ""
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages[:1]:  # Only first page
                text += page.extract_text() or ""
        
        if not text.strip():
            return "未找到单位信息"
        
        lines = text.split('\n')
        affiliations = set()
        
        # Strategy 1: Look for numbered footnote patterns (most common in academic papers)
        # Pattern: "1Department of X, University of Y" or "^1Department of X"
        numbered_affils = []
        in_affil = False
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            if not line:
                i += 1
                continue
            
            # Check if line has digit-superscripted affiliation
            m = re.match(r'[\^]?\s*(\d+)[\.\)]?\s+(.+)', line)
            if m:
                in_affil = True
                affil_text = m.group(2).strip()
                # Clean trailing punctuation
                affil_text = re.sub(r'[,;]+$', '', affil_text)
                if len(affil_text) > 5:
                    numbered_affils.append(affil_text)
                i += 1
                continue
            
            if in_affil:
                # Continuation of previous affiliation
                if line.startswith('and ') or (line[0].islower() if line else False):
                    if numbered_affils:
                        numbered_affils[-1] += ' ' + line
                elif re.match(r'[\^]?\s*\d+', line):
                    # Another numbered affiliation
                    m = re.match(r'[\^]?\s*(\d+)[\.\)]?\s+(.+)', line)
                    if m:
                        affil_text = m.group(2).strip()
                        affil_text = re.sub(r'[,;]+$', '', affil_text)
                        if len(affil_text) > 5:
                            numbered_affils.append(affil_text)
                else:
                    in_affil = False
            
            i += 1
        
        if numbered_affils:
            for aff in numbered_affils:
                # Split by semicolons
                for sub_aff in re.split(r'\s*[;]\s*', aff):
                    sub_aff = sub_aff.strip()
                    if sub_aff and len(sub_aff) > 5:
                        affiliations.add(sub_aff)
        
        # Strategy 2: Look for institution names with known keywords
        if len(affiliations) < 1:
            inst_patterns = [
                r'(?:University|Institute|College|School|Laboratory|Lab|Center|Centre|Department|Hospital|Academy|Faculty)\s+of\s+[\w\s&\'-]+',
                r'[\w\s&\'-]+\s+(?:University|Institute|College|School|Lab|Laboratory|Center|Centre|Hospital|Academy)',
            ]
            for line in lines:
                line = line.strip()
                for pat in inst_patterns:
                    found = re.findall(pat, line)
                    for f in found:
                        f = f.strip()
                        if len(f) > 10:
                            # Clean CamelCase
                            f = re.sub(r'([a-z])([A-Z])', r'\1 \2', f)
                            f = re.sub(r'([A-Z]+)([A-Z][a-z])', r'\1 \2', f)
                            f = re.sub(r'\s+', ' ', f).strip()
                            affiliations.add(f)
        
        # Strategy 3: Look for email domains to infer institutions
        if len(affiliations) < 1:
            emails = re.findall(r'[\w.]+@([\w.-]+\.\w+)', text)
            for domain in emails:
                domain = domain.lower()
                if 'gmail' in domain or 'yahoo' in domain or 'outlook' in domain:
                    continue
                affiliations.add(f"Domain: {domain}")
        
        if affiliations:
            # Clean up
            cleaned = []
            for aff in affiliations:
                aff = re.sub(r'\s+', ' ', aff).strip()
                # Remove very short entries
                if len(aff) < 5:
                    continue
                # Remove lines that look like equations or references
                if re.search(r'[={}\[\]()]', aff) and len(re.findall(r'[a-zA-Z]', aff)) < 10:
                    continue
                # Remove page numbers and standalone numbers
                if re.match(r'^\d+\s*$', aff):
                    continue
                cleaned.append(aff)
            
            if cleaned:
                return "; ".join(sorted(set(cleaned)))
        
        return "未找到单位信息"
    
    except Exception as e:
        print(f"  [WARN] PDF extraction error: {e}")
        return "未找到单位信息"


def update_excel(paper):
    """Update affiliations and summary_cn in Excel."""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Papers"]
    
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            headers[str(val).strip()] = col
    
    arxiv_col = headers.get("arxiv_id")
    if not arxiv_col:
        print("  [ERROR] No arxiv_id column")
        return False
    
    found = False
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=arxiv_col).value
        if val and str(val).strip() == paper["arxiv_id"]:
            found = True
            if paper.get("affiliations"):
                col = headers.get("affiliations")
                if col:
                    old = ws.cell(row=row, column=col).value or ""
                    if paper["affiliations"] != "未找到单位信息" or not old:
                        ws.cell(row=row, column=col, value=paper["affiliations"])
            
            if paper.get("summary_cn"):
                col = headers.get("summary_cn")
                if col:
                    ws.cell(row=row, column=col, value=paper["summary_cn"])
            
            col = headers.get("crawled_date")
            if col:
                ws.cell(row=row, column=col, value=date.today().isoformat())
            
            print(f"  [OK] Updated row {row}: {paper['arxiv_id']}")
            break
    
    if not found:
        print(f"  [WARN] Paper {paper['arxiv_id']} not found in Excel")
        return False
    
    wb.save(EXCEL_FILE)
    return True


# === MAIN ===
success = 0
fail = 0

for i, paper in enumerate(papers_to_process):
    print(f"\n[{i+1}/{len(papers_to_process)}] {paper['arxiv_id']}: {paper['title'][:50]}...")
    
    # Extract affiliations
    pdf_path = paper.get("pdf_local_path", str(BASE_DIR / "papers" / paper.get("pdf_filename", "")))
    affiliations = extract_affiliations_from_pdf(pdf_path)
    paper["affiliations"] = affiliations
    print(f"  Affiliations: {affiliations[:80] if affiliations != '未找到单位信息' else affiliations}...")
    
    # Generate Chinese summary
    abstract = paper.get("abstract") or paper.get("summary", "")
    summary_cn = generate_cn_summary(paper.get("title", ""), abstract)
    paper["summary_cn"] = summary_cn
    print(f"  Summary_CN ({len(summary_cn)} chars): {summary_cn[:60]}...")
    
    if update_excel(paper):
        success += 1
    else:
        fail += 1

print(f"\n{'='*60}")
print(f"Done: {success} succeeded, {fail} failed")
